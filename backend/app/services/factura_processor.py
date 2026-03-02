"""
Servicio de Procesamiento de Facturas.
Fase 3.5 - Carga de facturas, extracción de códigos, carga manual y desde CSV/Excel.
"""

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import BinaryIO, Dict, List, Optional, Set, Union

from loguru import logger
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.documento import Documento
from app.schemas.documento import EstadoDocumento, TipoDocumento
from app.services.code_extractor import ExtractorCodigos, TipoCodigo
from app.services.ocr_extractor import (
    CalidadImagen,
    ConfiguracionOCR,
    IdiomaOCR,
    OCRExtractor,
)
from app.services.pdf_extractor import PDFExtractor

try:
    from openpyxl import load_workbook
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False
    logger.warning("openpyxl no está instalado para lectura de Excel")


class OrigenCodigo(str, Enum):
    """Origen de los códigos de factura."""
    
    PDF_EXTRAIDO = "pdf_extraido"
    MANUAL = "manual"
    CSV = "csv"
    EXCEL = "excel"
    TEXTO = "texto"


@dataclass
class CodigoFactura:
    """Representa un código de factura con su origen."""
    
    codigo_original: str
    codigo_normalizado: str
    origen: OrigenCodigo
    linea_archivo: Optional[int] = None  # Línea en CSV/Excel
    columna_archivo: Optional[str] = None  # Columna en Excel
    pagina_pdf: Optional[int] = None  # Si viene de PDF
    valido: bool = True
    error: Optional[str] = None


@dataclass
class ResultadoCargaCodigos:
    """Resultado de la carga de códigos desde cualquier fuente."""
    
    exito: bool
    codigos: List[CodigoFactura] = field(default_factory=list)
    total_leidos: int = 0
    total_validos: int = 0
    total_duplicados: int = 0
    total_invalidos: int = 0
    duplicados: List[str] = field(default_factory=list)
    invalidos: List[str] = field(default_factory=list)
    origen: OrigenCodigo = OrigenCodigo.MANUAL
    mensaje: str = ""
    errores: List[str] = field(default_factory=list)


@dataclass
class ResumenFactura:
    """Resumen del procesamiento de una factura."""
    
    documento_id: Optional[int] = None
    nombre_archivo: str = ""
    tipo: TipoDocumento = TipoDocumento.FACTURA
    
    # Códigos
    total_codigos: int = 0
    codigos_unicos: int = 0
    codigos_por_origen: Dict[str, int] = field(default_factory=dict)
    
    # Estado
    exito: bool = False
    errores: List[str] = field(default_factory=list)
    
    # Tiempos
    tiempo_proceso_ms: int = 0
    fecha_proceso: Optional[datetime] = None
    
    def to_dict(self) -> dict:
        """Convierte a diccionario."""
        return {
            "documento_id": self.documento_id,
            "nombre_archivo": self.nombre_archivo,
            "tipo": self.tipo.value,
            "codigos": {
                "total": self.total_codigos,
                "unicos": self.codigos_unicos,
                "por_origen": self.codigos_por_origen,
            },
            "exito": self.exito,
            "errores": self.errores,
            "tiempo_proceso_ms": self.tiempo_proceso_ms,
            "fecha_proceso": self.fecha_proceso.isoformat() if self.fecha_proceso else None,
        }


class FacturaProcessor:
    """
    Procesador de facturas.
    
    Permite:
    - Cargar códigos desde PDF de factura
    - Cargar códigos manualmente (lista)
    - Cargar códigos desde archivo CSV
    - Cargar códigos desde archivo Excel
    - Normalizar todos los códigos
    """
    
    # Patrón para validar códigos (al menos 4 caracteres alfanuméricos)
    PATRON_CODIGO_VALIDO = re.compile(r'^[A-Z0-9\-\s]{4,}$', re.IGNORECASE)
    
    def __init__(
        self,
        db: Optional[Session] = None,
        usar_ocr: bool = True,
        idioma_ocr: IdiomaOCR = IdiomaOCR.ESPAÑOL_INGLES,
    ):
        """
        Inicializa el procesador de facturas.
        
        Args:
            db: Sesión de SQLAlchemy (opcional)
            usar_ocr: Usar OCR si el PDF es escaneado
            idioma_ocr: Idioma para OCR
        """
        self.db = db
        self.usar_ocr = usar_ocr
        
        # Inicializar extractores
        self.pdf_extractor = PDFExtractor()
        self.code_extractor = ExtractorCodigos()
        
        if usar_ocr:
            self.ocr_config = ConfiguracionOCR(
                idioma=idioma_ocr,
                calidad=CalidadImagen.ALTA,
            )
            self.ocr_extractor = OCRExtractor(self.ocr_config)
        else:
            self.ocr_extractor = None
        
        logger.info("FacturaProcessor inicializado")
    
    def normalizar_codigo(self, codigo: str) -> str:
        """
        Normaliza un código de factura.
        
        Args:
            codigo: Código original
            
        Returns:
            Código normalizado (mayúsculas, sin espacios/guiones)
        """
        if not codigo:
            return ""
        
        # Mayúsculas
        normalizado = codigo.upper().strip()
        
        # Eliminar espacios, guiones, puntos, guiones bajos
        normalizado = re.sub(r'[\s\-_\.]+', '', normalizado)
        
        return normalizado
    
    def validar_codigo(self, codigo: str) -> bool:
        """
        Valida si un código tiene formato aceptable.
        
        Args:
            codigo: Código a validar
            
        Returns:
            True si el código es válido
        """
        if not codigo or len(codigo.strip()) < 4:
            return False
        
        # Debe tener al menos un número o ser alfanumérico significativo
        return bool(self.PATRON_CODIGO_VALIDO.match(codigo.strip()))
    
    def cargar_desde_pdf(
        self,
        source: Union[str, Path, BinaryIO, bytes],
        nombre_archivo: str = "factura.pdf",
        tipos_codigo: Optional[List[TipoCodigo]] = None,
    ) -> ResultadoCargaCodigos:
        """
        Carga códigos desde un PDF de factura.
        
        Args:
            source: Fuente del PDF
            nombre_archivo: Nombre del archivo
            tipos_codigo: Tipos de código a extraer
            
        Returns:
            ResultadoCargaCodigos
        """
        import time
        inicio = time.time()
        
        resultado = ResultadoCargaCodigos(
            exito=False,
            origen=OrigenCodigo.PDF_EXTRAIDO,
        )
        
        try:
            # Obtener contenido
            contenido = self._obtener_contenido(source)
            
            # Extraer texto del PDF
            resultado_pdf = self.pdf_extractor.extraer(contenido)
            
            if not resultado_pdf.exito:
                resultado.errores.append(f"Error leyendo PDF: {resultado_pdf.error}")
                return resultado
            
            # Construir textos por página
            textos_paginas: Dict[int, str] = {}
            
            for pagina in resultado_pdf.paginas:
                if pagina.tiene_texto:
                    textos_paginas[pagina.numero] = pagina.texto
                elif self.usar_ocr and self.ocr_extractor and pagina.es_escaneado:
                    # Aplicar OCR
                    resultado_ocr = self.ocr_extractor.extraer_pagina_especifica(
                        contenido,
                        pagina.numero,
                    )
                    if resultado_ocr.exito and resultado_ocr.texto:
                        textos_paginas[pagina.numero] = resultado_ocr.texto
            
            if not textos_paginas:
                resultado.errores.append("No se pudo extraer texto del PDF")
                return resultado
            
            # Extraer códigos
            resultado_codigos = self.code_extractor.extraer_de_paginas(
                textos_paginas,
                tipos_filtro=tipos_codigo,
            )
            
            # Convertir a CodigoFactura
            codigos_vistos: Set[str] = set()
            
            for codigo_ext in resultado_codigos.codigos:
                normalizado = self.normalizar_codigo(codigo_ext.valor_original)
                
                if normalizado in codigos_vistos:
                    resultado.duplicados.append(codigo_ext.valor_original)
                    resultado.total_duplicados += 1
                    continue
                
                codigos_vistos.add(normalizado)
                
                codigo_factura = CodigoFactura(
                    codigo_original=codigo_ext.valor_original,
                    codigo_normalizado=normalizado,
                    origen=OrigenCodigo.PDF_EXTRAIDO,
                    pagina_pdf=codigo_ext.pagina,
                    valido=self.validar_codigo(normalizado),
                )
                
                resultado.codigos.append(codigo_factura)
                resultado.total_leidos += 1
                
                if codigo_factura.valido:
                    resultado.total_validos += 1
                else:
                    resultado.total_invalidos += 1
                    resultado.invalidos.append(codigo_ext.valor_original)
            
            resultado.exito = True
            resultado.mensaje = (
                f"Extraídos {resultado.total_validos} códigos válidos "
                f"de {resultado_pdf.metadata.total_paginas} páginas"
            )
            
        except Exception as e:
            logger.exception(f"Error cargando códigos desde PDF: {e}")
            resultado.errores.append(str(e))
        
        return resultado
    
    def cargar_manual(
        self,
        codigos: List[str],
    ) -> ResultadoCargaCodigos:
        """
        Carga códigos ingresados manualmente.
        
        Args:
            codigos: Lista de códigos
            
        Returns:
            ResultadoCargaCodigos
        """
        resultado = ResultadoCargaCodigos(
            exito=True,
            origen=OrigenCodigo.MANUAL,
        )
        
        codigos_vistos: Set[str] = set()
        
        for i, codigo in enumerate(codigos, 1):
            codigo = codigo.strip()
            if not codigo:
                continue
            
            normalizado = self.normalizar_codigo(codigo)
            
            if normalizado in codigos_vistos:
                resultado.duplicados.append(codigo)
                resultado.total_duplicados += 1
                continue
            
            codigos_vistos.add(normalizado)
            resultado.total_leidos += 1
            
            valido = self.validar_codigo(codigo)
            
            codigo_factura = CodigoFactura(
                codigo_original=codigo,
                codigo_normalizado=normalizado,
                origen=OrigenCodigo.MANUAL,
                linea_archivo=i,
                valido=valido,
            )
            
            resultado.codigos.append(codigo_factura)
            
            if valido:
                resultado.total_validos += 1
            else:
                resultado.total_invalidos += 1
                resultado.invalidos.append(codigo)
        
        resultado.mensaje = (
            f"Cargados {resultado.total_validos} códigos válidos "
            f"de {resultado.total_leidos} ingresados"
        )
        
        return resultado
    
    def cargar_desde_texto(
        self,
        texto: str,
        separador: Optional[str] = None,
    ) -> ResultadoCargaCodigos:
        """
        Carga códigos desde un bloque de texto (uno por línea o separados).
        
        Args:
            texto: Texto con códigos
            separador: Separador personalizado (None = líneas)
            
        Returns:
            ResultadoCargaCodigos
        """
        if separador:
            codigos = [c.strip() for c in texto.split(separador) if c.strip()]
        else:
            # Separar por líneas y también por comas si están en una línea
            lineas = texto.strip().split('\n')
            codigos = []
            for linea in lineas:
                # Si la línea tiene comas, separar también por comas
                if ',' in linea:
                    codigos.extend([c.strip() for c in linea.split(',') if c.strip()])
                elif linea.strip():
                    codigos.append(linea.strip())
        
        resultado = self.cargar_manual(codigos)
        resultado.origen = OrigenCodigo.TEXTO
        
        return resultado
    
    def cargar_desde_csv(
        self,
        source: Union[str, Path, BinaryIO, bytes],
        columna: int = 0,
        tiene_encabezado: bool = True,
        delimitador: str = ',',
        encoding: str = 'utf-8',
    ) -> ResultadoCargaCodigos:
        """
        Carga códigos desde un archivo CSV.
        
        Args:
            source: Fuente del CSV
            columna: Índice de columna con los códigos (0-based)
            tiene_encabezado: Si la primera fila es encabezado
            delimitador: Delimitador del CSV
            encoding: Encoding del archivo
            
        Returns:
            ResultadoCargaCodigos
        """
        resultado = ResultadoCargaCodigos(
            exito=False,
            origen=OrigenCodigo.CSV,
        )
        
        try:
            # Obtener contenido como texto
            if isinstance(source, bytes):
                contenido = source.decode(encoding)
            elif isinstance(source, (str, Path)):
                with open(source, 'r', encoding=encoding) as f:
                    contenido = f.read()
            else:
                contenido = source.read()
                if isinstance(contenido, bytes):
                    contenido = contenido.decode(encoding)
            
            # Parsear CSV
            reader = csv.reader(io.StringIO(contenido), delimiter=delimitador)
            
            codigos_vistos: Set[str] = set()
            linea_inicio = 1 if tiene_encabezado else 0
            
            for i, fila in enumerate(reader):
                if i < linea_inicio:
                    continue
                
                if len(fila) <= columna:
                    continue
                
                codigo = fila[columna].strip()
                if not codigo:
                    continue
                
                normalizado = self.normalizar_codigo(codigo)
                
                if normalizado in codigos_vistos:
                    resultado.duplicados.append(codigo)
                    resultado.total_duplicados += 1
                    continue
                
                codigos_vistos.add(normalizado)
                resultado.total_leidos += 1
                
                valido = self.validar_codigo(codigo)
                
                codigo_factura = CodigoFactura(
                    codigo_original=codigo,
                    codigo_normalizado=normalizado,
                    origen=OrigenCodigo.CSV,
                    linea_archivo=i + 1,
                    valido=valido,
                )
                
                resultado.codigos.append(codigo_factura)
                
                if valido:
                    resultado.total_validos += 1
                else:
                    resultado.total_invalidos += 1
                    resultado.invalidos.append(codigo)
            
            resultado.exito = True
            resultado.mensaje = (
                f"Cargados {resultado.total_validos} códigos válidos desde CSV"
            )
            
        except Exception as e:
            logger.exception(f"Error leyendo CSV: {e}")
            resultado.errores.append(f"Error leyendo CSV: {e}")
        
        return resultado
    
    def cargar_desde_excel(
        self,
        source: Union[str, Path, BinaryIO, bytes],
        hoja: Union[str, int] = 0,
        columna: Union[str, int] = 'A',
        fila_inicio: int = 2,
    ) -> ResultadoCargaCodigos:
        """
        Carga códigos desde un archivo Excel.
        
        Args:
            source: Fuente del Excel
            hoja: Nombre o índice de la hoja (0-based)
            columna: Letra o índice de columna
            fila_inicio: Fila de inicio (1-based, típicamente 2 si hay encabezado)
            
        Returns:
            ResultadoCargaCodigos
        """
        if not OPENPYXL_DISPONIBLE:
            return ResultadoCargaCodigos(
                exito=False,
                origen=OrigenCodigo.EXCEL,
                errores=["openpyxl no está instalado. Ejecuta: pip install openpyxl"],
            )
        
        resultado = ResultadoCargaCodigos(
            exito=False,
            origen=OrigenCodigo.EXCEL,
        )
        
        try:
            # Cargar workbook
            if isinstance(source, bytes):
                wb = load_workbook(io.BytesIO(source), read_only=True, data_only=True)
            elif isinstance(source, (str, Path)):
                wb = load_workbook(source, read_only=True, data_only=True)
            else:
                content = source.read()
                wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            
            # Seleccionar hoja
            if isinstance(hoja, int):
                ws = wb.worksheets[hoja]
            else:
                ws = wb[hoja]
            
            # Convertir columna a letra si es número
            if isinstance(columna, int):
                from openpyxl.utils import get_column_letter
                columna_letra = get_column_letter(columna + 1)
            else:
                columna_letra = columna.upper()
            
            codigos_vistos: Set[str] = set()
            
            # Iterar filas
            for i, fila in enumerate(ws.iter_rows(min_row=fila_inicio), start=fila_inicio):
                # Obtener celda de la columna
                col_idx = ord(columna_letra) - ord('A')
                if col_idx >= len(fila):
                    continue
                
                celda = fila[col_idx]
                valor = celda.value
                
                if valor is None:
                    continue
                
                codigo = str(valor).strip()
                if not codigo:
                    continue
                
                normalizado = self.normalizar_codigo(codigo)
                
                if normalizado in codigos_vistos:
                    resultado.duplicados.append(codigo)
                    resultado.total_duplicados += 1
                    continue
                
                codigos_vistos.add(normalizado)
                resultado.total_leidos += 1
                
                valido = self.validar_codigo(codigo)
                
                codigo_factura = CodigoFactura(
                    codigo_original=codigo,
                    codigo_normalizado=normalizado,
                    origen=OrigenCodigo.EXCEL,
                    linea_archivo=i,
                    columna_archivo=columna_letra,
                    valido=valido,
                )
                
                resultado.codigos.append(codigo_factura)
                
                if valido:
                    resultado.total_validos += 1
                else:
                    resultado.total_invalidos += 1
                    resultado.invalidos.append(codigo)
            
            wb.close()
            
            resultado.exito = True
            resultado.mensaje = (
                f"Cargados {resultado.total_validos} códigos válidos desde Excel"
            )
            
        except Exception as e:
            logger.exception(f"Error leyendo Excel: {e}")
            resultado.errores.append(f"Error leyendo Excel: {e}")
        
        return resultado
    
    def procesar_factura(
        self,
        source: Union[str, Path, BinaryIO, bytes],
        nombre_archivo: str,
        usuario_id: Optional[int] = None,
        descripcion: Optional[str] = None,
    ) -> ResumenFactura:
        """
        Procesa una factura PDF completa.
        
        Args:
            source: Fuente del PDF
            nombre_archivo: Nombre del archivo
            usuario_id: ID del usuario
            descripcion: Descripción opcional
            
        Returns:
            ResumenFactura
        """
        import time
        inicio = time.time()
        
        resumen = ResumenFactura(
            nombre_archivo=nombre_archivo,
            tipo=TipoDocumento.FACTURA,
            fecha_proceso=datetime.now(),
        )
        
        try:
            # Cargar códigos desde PDF
            resultado = self.cargar_desde_pdf(source, nombre_archivo)
            
            if not resultado.exito:
                resumen.errores.extend(resultado.errores)
                return resumen
            
            resumen.total_codigos = resultado.total_leidos
            resumen.codigos_unicos = resultado.total_validos
            resumen.codigos_por_origen = {OrigenCodigo.PDF_EXTRAIDO.value: resultado.total_validos}
            
            # Almacenar en DB si está disponible
            if self.db:
                contenido = self._obtener_contenido(source)
                documento = self._almacenar_factura(
                    contenido=contenido,
                    nombre_archivo=nombre_archivo,
                    codigos=resultado.codigos,
                    usuario_id=usuario_id,
                    descripcion=descripcion,
                )
                resumen.documento_id = documento.id
            
            resumen.exito = True
            
        except Exception as e:
            logger.exception(f"Error procesando factura: {e}")
            resumen.errores.append(str(e))
        
        finally:
            resumen.tiempo_proceso_ms = int((time.time() - inicio) * 1000)
        
        return resumen
    
    def _obtener_contenido(
        self,
        source: Union[str, Path, BinaryIO, bytes],
    ) -> bytes:
        """Obtiene bytes del archivo."""
        if isinstance(source, bytes):
            return source
        elif isinstance(source, (str, Path)):
            with open(source, "rb") as f:
                return f.read()
        else:
            return source.read()
    
    def _almacenar_factura(
        self,
        contenido: bytes,
        nombre_archivo: str,
        codigos: List[CodigoFactura],
        usuario_id: Optional[int],
        descripcion: Optional[str],
    ) -> Documento:
        """Almacena la factura en la base de datos."""
        import hashlib
        import uuid
        
        unique_name = f"{uuid.uuid4()}_{nombre_archivo}"
        storage_path = f"{settings.UPLOAD_DIR}/facturas/{unique_name}"
        checksum = hashlib.sha256(contenido).hexdigest()
        
        documento = Documento(
            nombre=unique_name,
            nombre_original=nombre_archivo,
            tipo=TipoDocumento.FACTURA,
            estado=EstadoDocumento.PROCESADO,
            tamaño=len(contenido),
            descripcion=descripcion,
            storage_path=storage_path,
            content_type="application/pdf",
            checksum=checksum,
            texto_extraido=True,
            usuario_id=usuario_id,
        )
        
        self.db.add(documento)
        self.db.flush()
        
        # Guardar archivo
        Path(storage_path).parent.mkdir(parents=True, exist_ok=True)
        with open(storage_path, "wb") as f:
            f.write(contenido)
        
        self.db.commit()
        
        return documento
    
    def obtener_codigos_normalizados(
        self,
        resultado: ResultadoCargaCodigos,
        solo_validos: bool = True,
    ) -> List[str]:
        """
        Obtiene lista simple de códigos normalizados.
        
        Args:
            resultado: Resultado de carga de códigos
            solo_validos: Solo retornar códigos válidos
            
        Returns:
            Lista de códigos normalizados
        """
        if solo_validos:
            return [c.codigo_normalizado for c in resultado.codigos if c.valido]
        return [c.codigo_normalizado for c in resultado.codigos]


# Instancia global
factura_processor = FacturaProcessor()


# Funciones de conveniencia
def cargar_codigos_desde_texto(texto: str) -> ResultadoCargaCodigos:
    """Carga códigos desde un bloque de texto."""
    return factura_processor.cargar_desde_texto(texto)


def cargar_codigos_desde_csv(
    source: Union[str, Path, BinaryIO, bytes],
    columna: int = 0,
) -> ResultadoCargaCodigos:
    """Carga códigos desde CSV."""
    return factura_processor.cargar_desde_csv(source, columna=columna)


def cargar_codigos_desde_excel(
    source: Union[str, Path, BinaryIO, bytes],
    columna: str = 'A',
) -> ResultadoCargaCodigos:
    """Carga códigos desde Excel."""
    return factura_processor.cargar_desde_excel(source, columna=columna)
